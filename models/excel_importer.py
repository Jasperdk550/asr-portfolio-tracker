"""
Excel Importer
--------------
Reads a user's existing portfolio from an Excel (.xlsx / .xls) or CSV file
and imports it into the tracker's portfolio.json.

Design goals
------------
- Tolerant:   accepts many real-world column-name variants (e.g. "Ticker",
              "Symbol", "ISIN-ticker", "Stock") via fuzzy matching so the
              user does not need to rename their spreadsheet.
- Transparent: produces a detailed validation report before any data is
              written — the user reviews what will be imported.
- Safe:        never overwrites existing positions; appends only.
              Runs a dry-run by default; writes only when confirmed.

Expected columns (case-insensitive, flexible names)
----------------------------------------------------
  REQUIRED
    ticker         → Ticker, Symbol, Stock, ISIN, Code, Asset
    quantity       → Quantity, Qty, Units, Shares, Amount, Position, Holdings
    purchase_price → Price, Buy Price, Avg Price, Cost, Purchase Price,
                     Avg Cost, Entry Price, Cost Basis per Share

  OPTIONAL  (sensible defaults applied if missing)
    sector         → Sector, Industry, Category          [default: "Unknown"]
    asset_class    → Asset Class, Class, Type, Kind      [default: "Equity"]
    purchase_date  → Date, Buy Date, Purchase Date,
                     Trade Date, Entry Date              [default: today]
    name           → Name, Description, Company, Full Name
    currency       → Currency, CCY                       [default: "USD"]

Template
--------
Run  python main.py import-excel --create-template
to generate portfolio_template.xlsx with the correct headers and
three example rows.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import pandas as pd
import numpy as np

from models.asset import Asset, VALID_ASSET_CLASSES


# ---------------------------------------------------------------------------
# Column synonym map
# ---------------------------------------------------------------------------

COLUMN_SYNONYMS: Dict[str, List[str]] = {
    "ticker": [
        "ticker", "symbol", "stock", "isin", "code", "asset",
        "instrument", "security", "ric", "cusip", "asset_ticker",
        "stock_symbol", "share", "equity",
    ],
    "quantity": [
        "quantity", "qty", "units", "shares", "amount", "position",
        "holdings", "num_shares", "number_of_shares", "volume",
        "lot_size", "size", "number", "shares_held",
    ],
    "purchase_price": [
        "purchase_price", "price", "buy_price", "avg_price", "cost",
        "average_price", "entry_price", "cost_basis", "avg_cost",
        "average_cost", "purchase", "paid", "cost_per_share",
        "avg_purchase_price", "book_price", "book_value_per_share",
    ],
    "sector": [
        "sector", "industry", "category", "gics_sector",
        "sector_name", "business_sector",
    ],
    "asset_class": [
        "asset_class", "class", "type", "kind", "asset_type",
        "instrument_type", "security_type", "product_type",
    ],
    "purchase_date": [
        "purchase_date", "date", "buy_date", "trade_date",
        "entry_date", "acquisition_date", "transaction_date",
        "open_date", "start_date", "date_purchased",
    ],
    "name": [
        "name", "description", "company", "full_name",
        "company_name", "security_name", "instrument_name",
        "long_name", "asset_name",
    ],
    "currency": [
        "currency", "ccy", "cur", "fx", "denomination",
    ],
}

# Default values when a column is missing
DEFAULTS = {
    "sector":        "Unknown",
    "asset_class":   "Equity",
    "purchase_date": str(date.today()),
    "name":          "",
    "currency":      "USD",
}


# ---------------------------------------------------------------------------
# Result dataclasses
# ---------------------------------------------------------------------------

@dataclass
class ImportRow:
    """One successfully parsed row from the spreadsheet."""
    ticker:         str
    sector:         str
    asset_class:    str
    quantity:       float
    purchase_price: float
    purchase_date:  str
    name:           str
    currency:       str
    source_row:     int    # 1-based row number in the original file


@dataclass
class ImportError:
    """A validation error on a specific row."""
    row:     int
    column:  str
    value:   str
    reason:  str


@dataclass
class ImportResult:
    """Full output of :func:`parse_excel`."""
    filepath:       str
    sheet_name:     str
    total_rows:     int
    valid_rows:     List[ImportRow]
    errors:         List[ImportError]
    column_mapping: Dict[str, str]    # canonical_name -> actual_col_header
    warnings:       List[str]

    @property
    def n_valid(self) -> int:
        return len(self.valid_rows)

    @property
    def n_errors(self) -> int:
        return len(self.errors)

    @property
    def has_errors(self) -> bool:
        return len(self.errors) > 0


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def parse_excel(
    filepath: str,
    sheet_name: Optional[str] = None,
) -> ImportResult:
    """
    Read an Excel or CSV file and return a validated :class:`ImportResult`.

    Parameters
    ----------
    filepath   : Path to .xlsx, .xls or .csv file.
    sheet_name : Specific sheet to read.  If None, uses the first sheet.

    Returns
    -------
    ImportResult (never raises; all problems go into result.errors)
    """
    path = Path(filepath)
    warnings: List[str] = []

    # ── Load file ────────────────────────────────────────────────────
    try:
        if path.suffix.lower() == ".csv":
            df = pd.read_csv(filepath, dtype=str)
            used_sheet = "CSV"
        else:
            xl = pd.ExcelFile(filepath)
            if sheet_name is None:
                # Pick the first sheet that is not obviously metadata
                candidates = [s for s in xl.sheet_names
                              if not re.search(r'(readme|info|legend|meta|about|instr)',
                                               s, re.I)]
                sheet_name = candidates[0] if candidates else xl.sheet_names[0]
                if len(xl.sheet_names) > 1:
                    warnings.append(
                        f"Multiple sheets found: {xl.sheet_names}. "
                        f"Using '{sheet_name}'. Pass --sheet to override."
                    )
            df = pd.read_excel(filepath, sheet_name=sheet_name, dtype=str)
            used_sheet = sheet_name
    except Exception as exc:
        return ImportResult(
            filepath=str(filepath),
            sheet_name=str(sheet_name or ""),
            total_rows=0,
            valid_rows=[],
            errors=[ImportError(0, "file", str(filepath), str(exc))],
            column_mapping={},
            warnings=[],
        )

    # ── Drop completely empty rows/columns ───────────────────────────
    df = df.dropna(how="all").reset_index(drop=True)
    if df.empty:
        return ImportResult(
            filepath=str(filepath), sheet_name=used_sheet,
            total_rows=0, valid_rows=[],
            errors=[ImportError(0, "sheet", used_sheet, "Sheet is empty.")],
            column_mapping={}, warnings=warnings,
        )

    # ── Map columns ──────────────────────────────────────────────────
    col_map = _map_columns(df.columns.tolist())

    # Check required columns found
    errors: List[ImportError] = []
    for required in ("ticker", "quantity", "purchase_price"):
        if required not in col_map:
            errors.append(ImportError(
                0, required, "",
                f"Required column '{required}' not found. "
                f"Available columns: {list(df.columns)}"
            ))
    if errors:
        return ImportResult(
            filepath=str(filepath), sheet_name=used_sheet,
            total_rows=len(df), valid_rows=[], errors=errors,
            column_mapping=col_map, warnings=warnings,
        )

    # ── Parse rows ───────────────────────────────────────────────────
    valid_rows: List[ImportRow] = []

    for idx, row in df.iterrows():
        row_num = int(idx) + 2  # +2: 1 for 1-based, 1 for header
        row_errors: List[ImportError] = []

        def get(canonical: str) -> str:
            if canonical in col_map:
                val = row.get(col_map[canonical], "")
                return "" if pd.isna(val) else str(val).strip()
            return DEFAULTS.get(canonical, "")

        # ── ticker ───────────────────────────────────────────────────
        ticker_raw = get("ticker")
        if not ticker_raw:
            row_errors.append(ImportError(row_num, "ticker", "", "Ticker is empty."))
        ticker = ticker_raw.upper().replace(" ", "-")

        # ── quantity ─────────────────────────────────────────────────
        qty_raw = get("quantity")
        quantity = _parse_number(qty_raw)
        if quantity is None or quantity <= 0:
            row_errors.append(ImportError(
                row_num, "quantity", qty_raw,
                "Quantity must be a positive number."
            ))

        # ── purchase_price ────────────────────────────────────────────
        price_raw = get("purchase_price")
        price = _parse_number(price_raw)
        if price is None or price <= 0:
            row_errors.append(ImportError(
                row_num, "purchase_price", price_raw,
                "Purchase price must be a positive number."
            ))

        # ── asset_class ──────────────────────────────────────────────
        ac_raw = get("asset_class")
        asset_class = _match_asset_class(ac_raw) if ac_raw else DEFAULTS["asset_class"]

        # ── purchase_date ─────────────────────────────────────────────
        date_raw = get("purchase_date")
        purchase_date = _parse_date(date_raw) if date_raw else DEFAULTS["purchase_date"]
        if purchase_date is None:
            warnings.append(
                f"Row {row_num}: Could not parse date '{date_raw}', "
                f"using today ({DEFAULTS['purchase_date']})."
            )
            purchase_date = DEFAULTS["purchase_date"]

        if row_errors:
            errors.extend(row_errors)
            continue

        valid_rows.append(ImportRow(
            ticker         = ticker,
            sector         = get("sector") or DEFAULTS["sector"],
            asset_class    = asset_class,
            quantity       = float(quantity),
            purchase_price = float(price),
            purchase_date  = purchase_date,
            name           = get("name") or "",
            currency       = (get("currency") or DEFAULTS["currency"]).upper(),
            source_row     = row_num,
        ))

    return ImportResult(
        filepath       = str(filepath),
        sheet_name     = used_sheet,
        total_rows     = len(df),
        valid_rows     = valid_rows,
        errors         = errors,
        column_mapping = col_map,
        warnings       = warnings,
    )


def create_template(filepath: str = "portfolio_template.xlsx") -> str:
    """
    Generate a ready-to-fill Excel template with correct headers,
    column descriptions, and three example rows.
    """
    from openpyxl import Workbook
    from openpyxl.styles import (Font, PatternFill, Alignment,
                                  Border, Side, GradientFill)
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ── Sheet 1: Portfolio ───────────────────────────────────────────
    ws = wb.active
    ws.title = "Portfolio"

    # Colour palette
    HEADER_BG   = "003082"   # a.s.r. blue
    HEADER_FG   = "FFFFFF"
    EXAMPLE_BG  = "EEF2FF"
    REQUIRED_BG = "FFF3CD"
    ALT_BG      = "F8F9FF"
    BORDER_COL  = "CCCCCC"

    thin = Side(style="thin", color=BORDER_COL)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = [
        ("Ticker",         "REQUIRED — e.g. AAPL, MSFT, BTC-USD",   True),
        ("Sector",         "e.g. Technology, Healthcare, Energy",     False),
        ("Asset Class",    "Equity / ETF / Bond / Crypto / Commodity / Real Estate / Other", False),
        ("Quantity",       "REQUIRED — number of units / shares",     True),
        ("Purchase Price", "REQUIRED — price per unit at purchase",   True),
        ("Purchase Date",  "Format: YYYY-MM-DD  (e.g. 2023-06-15)",  False),
        ("Name",           "Optional — company or instrument name",   False),
        ("Currency",       "Optional — e.g. USD, EUR, GBP",          False),
    ]

    col_widths = [14, 18, 22, 12, 16, 16, 28, 10]

    # Header row
    for col_idx, (header, _, required) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = Font(bold=True, color=HEADER_FG, name="Arial", size=10)
        cell.fill      = PatternFill("solid", fgColor=HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = border
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths[col_idx - 1]

    ws.row_dimensions[1].height = 28

    # Description row (row 2 — light grey, italic)
    for col_idx, (_, desc, required) in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_idx, value=desc)
        cell.font      = Font(italic=True, color="666666", name="Arial", size=8)
        cell.fill      = PatternFill("solid", fgColor="F5F5F5")
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border    = border
    ws.row_dimensions[2].height = 32

    # Example rows (rows 3-5)
    examples = [
        ("AAPL",   "Technology",    "Equity",      "10",  "178.50", "2023-06-15", "Apple Inc.",           "USD"),
        ("MSFT",   "Technology",    "Equity",      "5",   "380.00", "2023-09-01", "Microsoft Corp.",      "USD"),
        ("BND",    "Fixed Income",  "Bond",        "50",  "72.30",  "2024-01-10", "Vanguard Bond ETF",    "USD"),
        ("GLD",    "Commodity",     "Commodity",   "8",   "185.00", "2024-03-20", "Gold ETF",             "USD"),
        ("BTC-USD","Crypto",        "Crypto",      "0.5", "42000",  "2024-02-01", "Bitcoin",              "USD"),
    ]

    for row_offset, example in enumerate(examples):
        row_num = row_offset + 3
        bg = EXAMPLE_BG if row_offset % 2 == 0 else ALT_BG
        for col_idx, value in enumerate(example, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=value)
            cell.font      = Font(name="Arial", size=10, color="333333")
            cell.fill      = PatternFill("solid", fgColor=bg)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border    = border

    # Freeze top 2 rows
    ws.freeze_panes = "A3"

    # ── Sheet 2: Instructions ────────────────────────────────────────
    wi = wb.create_sheet("Instructions")
    instructions = [
        ("a.s.r. Portfolio Tracker — Import Instructions", True, "003082", 14),
        ("", False, None, 11),
        ("HOW TO USE THIS TEMPLATE", True, "003082", 11),
        ("1.  Fill in your positions on the 'Portfolio' sheet starting from row 3.", False, None, 10),
        ("2.  Rows 1-2 contain headers and descriptions — do not delete them.", False, None, 10),
        ("3.  You may add or delete rows freely below row 2.", False, None, 10),
        ("4.  Save the file as .xlsx.", False, None, 10),
        ("5.  Run:  python main.py import-excel --file portfolio_template.xlsx", False, None, 10),
        ("", False, None, 10),
        ("REQUIRED COLUMNS", True, "C0392B", 11),
        ("  Ticker         — Stock exchange symbol (e.g. AAPL, MSFT, BTC-USD).", False, None, 10),
        ("  Quantity        — Number of shares / units purchased.", False, None, 10),
        ("  Purchase Price  — Price per unit at the time of purchase.", False, None, 10),
        ("", False, None, 10),
        ("OPTIONAL COLUMNS", True, "27AE60", 11),
        ("  Sector          — Business sector (default: Unknown).", False, None, 10),
        ("  Asset Class     — Equity / ETF / Bond / Crypto / Commodity / Real Estate / Other", False, None, 10),
        ("  Purchase Date   — Format YYYY-MM-DD (default: today).", False, None, 10),
        ("  Name            — Human-readable name (auto-fetched if blank).", False, None, 10),
        ("  Currency        — ISO currency code, e.g. USD, EUR (default: USD).", False, None, 10),
        ("", False, None, 10),
        ("FLEXIBLE COLUMN NAMES", True, "7D3C98", 11),
        ("  The importer accepts many variants — you do not need to rename your", False, None, 10),
        ("  existing spreadsheet. Examples that are auto-recognised:", False, None, 10),
        ("  'Symbol', 'Stock' → Ticker", False, None, 10),
        ("  'Shares', 'Units', 'Holdings' → Quantity", False, None, 10),
        ("  'Avg Price', 'Cost', 'Entry Price' → Purchase Price", False, None, 10),
        ("  'Buy Date', 'Trade Date' → Purchase Date", False, None, 10),
        ("", False, None, 10),
        ("MULTIPLE SHEETS", True, "2980B9", 11),
        ("  If your file has multiple sheets, use:  --sheet 'SheetName'", False, None, 10),
        ("  to specify which sheet to import.", False, None, 10),
    ]
    wi.column_dimensions["A"].width = 80
    for row_idx, (text, bold, color, size) in enumerate(instructions, 1):
        cell = wi.cell(row=row_idx, column=1, value=text)
        cell.font = Font(bold=bold,
                         color=color if color else "333333",
                         name="Arial", size=size)
        cell.alignment = Alignment(vertical="center")
        wi.row_dimensions[row_idx].height = 16 if text else 8

    Path(filepath).parent.mkdir(parents=True, exist_ok=True)
    wb.save(filepath)
    return filepath


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

def _normalise(s: str) -> str:
    """Lowercase, strip, collapse whitespace and punctuation to underscore."""
    s = s.lower().strip()
    s = re.sub(r"[\s\-/\\\.]+", "_", s)
    s = re.sub(r"[^a-z0-9_]", "", s)
    return s


def _map_columns(actual_columns: List[str]) -> Dict[str, str]:
    """
    Build a mapping from canonical field name → actual column header.
    Uses exact match first, then synonym match, then fuzzy substring match.
    """
    normalised = {_normalise(c): c for c in actual_columns}
    mapping: Dict[str, str] = {}

    for canonical, synonyms in COLUMN_SYNONYMS.items():
        if canonical in mapping:
            continue
        # Exact / normalised match against synonyms
        for syn in synonyms:
            norm_syn = _normalise(syn)
            if norm_syn in normalised:
                mapping[canonical] = normalised[norm_syn]
                break
        if canonical in mapping:
            continue
        # Substring match (e.g. "Avg. Purchase Price" contains "purchase_price")
        for norm_col, orig_col in normalised.items():
            for syn in synonyms:
                if _normalise(syn) in norm_col or norm_col in _normalise(syn):
                    mapping[canonical] = orig_col
                    break
            if canonical in mapping:
                break

    return mapping


def _parse_number(raw: str) -> Optional[float]:
    """Parse a numeric string, stripping currency symbols and commas."""
    if not raw:
        return None
    cleaned = re.sub(r"[€$£¥,\s]", "", str(raw))
    # Handle parentheses as negative: (1234) → -1234
    cleaned = re.sub(r"^\((.+)\)$", r"-\1", cleaned)
    try:
        return float(cleaned)
    except (ValueError, TypeError):
        return None


def _parse_date(raw: str) -> Optional[str]:
    """Try to parse a date string into ISO format YYYY-MM-DD."""
    if not raw:
        return None
    formats = [
        "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y",
        "%d-%m-%Y", "%m-%d-%Y", "%Y/%m/%d",
        "%d.%m.%Y", "%Y.%m.%d",
        "%d %b %Y", "%d %B %Y", "%b %d, %Y", "%B %d, %Y",
        "%Y%m%d",
    ]
    raw_clean = str(raw).strip()

    # pandas may parse Excel serial dates as floats
    try:
        serial = float(raw_clean)
        # Excel serial date: days since 1899-12-30
        parsed = pd.Timestamp("1899-12-30") + pd.Timedelta(days=serial)
        return parsed.strftime("%Y-%m-%d")
    except (ValueError, OverflowError):
        pass

    for fmt in formats:
        try:
            return datetime.strptime(raw_clean, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue

    # Let pandas try as last resort
    try:
        return pd.to_datetime(raw_clean, dayfirst=True).strftime("%Y-%m-%d")
    except Exception:
        return None


def _match_asset_class(raw: str) -> str:
    """Map a free-text asset class to a valid VALID_ASSET_CLASSES entry."""
    if not raw:
        return DEFAULTS["asset_class"]
    raw_lower = raw.lower()
    mapping = {
        "equity":      ["equity", "stock", "share", "equities", "common"],
        "etf":         ["etf", "fund", "exchange traded", "index fund", "mutual"],
        "bond":        ["bond", "fixed", "debt", "notes", "treasury", "gilt", "income"],
        "crypto":      ["crypto", "bitcoin", "eth", "digital", "coin", "token"],
        "commodity":   ["commodity", "commodit", "gold", "silver", "oil", "metal", "energy"],
        "real estate": ["real estate", "reit", "property", "realestate"],
        "other":       ["other", "alternative", "misc", "structured"],
    }
    for canonical, keywords in mapping.items():
        if any(kw in raw_lower for kw in keywords):
            # Match to the exact spelling in VALID_ASSET_CLASSES
            for valid in VALID_ASSET_CLASSES:
                if valid.lower() == canonical:
                    return valid
    return DEFAULTS["asset_class"]
